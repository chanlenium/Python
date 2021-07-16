import openpyxl
from selenium import webdriver

book = openpyxl.load_workbook('한국표준산업분류(10차)_크롤링.xlsx')
sheet = book.get_sheet_by_name("10차개정한국표준산업분류")
for r in sheet.rows: # 행 순회
    row_index = r[0].row   # 행 인덱스
    코드 = r[0].value
    #print(코드)
    #print(r[0].value)

    browser = webdriver.Chrome('/Users/biophdeo/PycharmProjects/Scraping/chromedriver')
    browser.get("https://kssc.kostat.go.kr:8443/ksscNew_web/link.do?gubun=001")
    browser.implicitly_wait(3)  # 웹브라우저와 똑같이 작동하기 때문에 서버에서 파일을 전부 읽어오기 위해 임의로 시간 조정을 해줌

    from selenium.common.exceptions import NoSuchElementException
    content = browser.find_element_by_xpath('//*[@id="ksscSearchCommonForm"]/div/table/tbody/tr/th[1]')
    차수 = '10' # '09', '10'중 택일
    from selenium.webdriver.support.ui import Select
    select = Select(browser.find_element_by_xpath('//*[@id="strCategoryDegree"]'))
    select.select_by_visible_text(차수)

    content = browser.find_element_by_xpath('//*[@id="ksscSearchCommonForm"]/div/table/tbody/tr/th[3]')
    구분 = '분류코드' # 전체, 분류코드, 색인어 중 택일
    select = Select(browser.find_element_by_xpath('//*[@id="strCategoryType"]'))
    select.select_by_visible_text(구분)

    content = browser.find_element_by_xpath('//*[@id="ksscSearchCommonForm"]/div/table/tbody/tr/th[7]')
    browser.find_element_by_xpath('//*[@id="strCategoryCodeName"]').send_keys(코드)
    browser.find_element_by_xpath('//*[@id="ksscSearchCommonForm"]/div/table/tbody/tr/th[9]/span/button').click()

    분류코드 = browser.find_element_by_xpath('//*[@id="contents"]/div[3]/table/tbody/tr[1]/td[1]/b').text
    print("분류코드 = ", 분류코드)
    분류명_한글 = browser.find_element_by_xpath('//*[@id="contents"]/div[3]/table/tbody/tr[1]/td[2]').text
    print("분류명(한글) = ", 분류명_한글)
    분류명_영문 = browser.find_element_by_xpath('//*[@id="contents"]/div[3]/table/tbody/tr[2]/td').text
    print("분류명(영문) = ", 분류명_영문)
    설명_한글 = browser.find_element_by_xpath('//*[@id="contents"]/div[3]/table/tbody/tr[4]/td').text
    print("설명(한글) = ", 설명_한글)
    색인어 = browser.find_element_by_xpath('//*[@id="contents"]/div[3]/table/tbody/tr[5]/td').text
    print("색인어 = ", 색인어)

    sheet.cell(row=row_index, column=1).value = 코드
    sheet.cell(row=row_index, column=2).value = 분류명_한글
    sheet.cell(row=row_index, column=3).value = 분류명_영문
    sheet.cell(row=row_index, column=4).value = 설명_한글
    sheet.cell(row=row_index, column=5).value = 색인어
    browser.close()

book.save("result.xlsx")
book.close()